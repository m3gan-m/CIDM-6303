""" Good references for more info: 
https://openpyxl.readthedocs.io/en/stable/index.html
http://zetcode.com/python/openpyxl/
"""
#megan moore
import openpyxl

wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)
sheet = wb["Sheet1"]
wb.create_sheet("Sheet2")

cell = sheet["a1"]
print(cell.value)
cell.value = "Transaction ID"
print(cell.row)
print(cell.column)
print(cell.coordinate)
cell = sheet.cell(row=1,column=1)

print("-"*30)
for c in range(1,sheet.max_column +1):
    for r in range(1, sheet.max_row +1):
        cell = sheet.cell(row=r, column=c)
        print(cell.value)

print("-"*30)

column = sheet["a"]
print(column)
for cell in column:
    print(cell.value)

print("-"*30)
columns = sheet["a:c"]
print(columns)
for column in columns:
    for cell in column:
        print(cell.value)

range = sheet["a1:c3"]
first_row=sheet[1]
three_rows=sheet[1:3]

data=[1004,4,9.99]
sheet.append(data)

new_rows_data =[
    [1004,4,9.99],
    [1005,5,7.52],
    [1006,6,1.99]
]
for row in new_rows_data:
    sheet.append(row)

wb.save("transactions2.xlsx")